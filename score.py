import openpyxl 
from konlpy.tag import Kkma
from konlpy.utils import pprint

kkma = Kkma()

result = openpyxl.Workbook()
resultSheet = result['Sheet']

test = openpyxl.load_workbook(filename='scoredata.xlsx')
sheet =  test.worksheets[0]
score = openpyxl.load_workbook(filename='score.xlsx')
scoreSheet = score.worksheets[0]
max_score_row = scoreSheet.max_row + 1
max_data_row = sheet.max_row + 1

indexArr = []

for i in range(3, max_data_row):
    if(sheet.cell(row = i, column = 1).value==None):
        max_data_row = i
        break
for i in range(2, max_score_row):
    if(scoreSheet.cell(row = i, column = 1).value==None):
        max_score_row = i
        break

for j in range(2,max_score_row):
    index = 3
    while True:
        if(scoreSheet.cell(j,index).value==None):
            indexArr.append(index)
            break
        else:
            index = index + 1

resultSheet.cell(row=1,column=1).value = "일렬 번호"
resultSheet.cell(row=1,column=2).value = "지원분야 대분류"
resultSheet.cell(row=1,column=3).value = "문항 구분"
resultSheet.cell(row=1,column=4).value = "문항명"
for i in range(2,max_score_row):
    resultSheet.cell(row=1,column=i+3).value = scoreSheet.cell(row=i,column=1).value

for i in range(3,max_data_row):
    resultSheet.cell(row=i-1,column=1).value = sheet.cell(row = i, column = 1).value
    resultSheet.cell(row=i-1,column=2).value = sheet.cell(row = i, column = 6).value
    resultSheet.cell(row=i-1,column=3).value = sheet.cell(row = i, column = 10).value
    resultSheet.cell(row=i-1,column=4).value = sheet.cell(row = i, column = 11).value
    tempResult = kkma.pos(sheet.cell(row = i, column = 12).value)
    for k in range(0,len(tempResult)):
        if(tempResult[k][1]=='VV'):
            sheet.cell(row = i, column = 12).value = str(sheet.cell(row = i, column = 12).value) + " " + str(tempResult[k][0]) +"다"
            continue
        elif(tempResult[k][1]=='VA'):
            sheet.cell(row = i, column = 12).value = str(sheet.cell(row = i, column = 12).value) + " " + str(tempResult[k][0]) + "다"
            continue
        elif(tempResult[k][1]=='NNG'):
            sheet.cell(row = i, column = 12).value = str(sheet.cell(row = i, column = 12).value) + " " + str(tempResult[k][0])
            continue
        elif(tempResult[k][1]=='MAG'):
            sheet.cell(row = i, column = 12).value = str(sheet.cell(row = i, column = 12).value) + " " + str(tempResult[k][0])
            continue
        elif(tempResult[k][1]=='XR'):
            sheet.cell(row = i, column = 12).value = str(sheet.cell(row = i, column = 12).value) + " " + str(tempResult[k][0])
            continue

    for j in range(0,len(indexArr)):
        resultSheet.cell(row=i-1,column=j+5).value = 0
        for k in range(3,indexArr[j]):
            resultSheet.cell(row=i-1,column=j+5).value = int(resultSheet.cell(row=i-1,column=j+5).value) + int(sheet.cell(row = i, column = 12).value.count(scoreSheet.cell(row=j+2, column = k).value))
        resultSheet.cell(row=i-1,column=j+5).value = int(resultSheet.cell(row=i-1,column=j+5).value) * int(scoreSheet.cell(row=j+2, column = 2).value)



result.save('./public/result_score.xlsx')
