import openpyxl 
from konlpy.tag import Kkma
from konlpy.utils import pprint
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('text', type=str,
                help="확인 하고싶은 단어")
args = parser.parse_args()

kkma = Kkma()

result = openpyxl.Workbook()
resultSheet = result['Sheet']

test = openpyxl.load_workbook(filename='data1.xlsx')
sheet = test.worksheets[0]
exception = openpyxl.load_workbook(filename='exception1.xlsx')
exceptionSheet = exception.worksheets[0]

max_one_exception_row = exceptionSheet.max_row + 1
max_data_row = sheet.max_row + 1

for i in range(2, max_data_row):
    if(sheet.cell(row = i, column = 1).value==None):
        max_data_row = i
        break
for i in range(2, max_one_exception_row):
    if(exceptionSheet.cell(row = i, column = 1).value==None):
        max_one_exception_row = i
        break

for i in range(2, max_data_row):
    tempResult = kkma.pos(sheet.cell(row = i, column = 1).value)
    resultSheet.cell(row=i,column=1).value=args.text
    for k in range(0,len(tempResult)):
        if(tempResult[k][1]=='VV'):
            resultSheet.cell(row=i,column=1).value = str(resultSheet.cell(row=i,column=1).value) + " " + str(tempResult[k][0]) +"다"
            continue
        elif(tempResult[k][1]=='VA'):
            resultSheet.cell(row=i,column=1).value = str(resultSheet.cell(row=i,column=1).value) + " " + str(tempResult[k][0]) + "다"
            continue
        elif(tempResult[k][1]=='NNG'):
            resultSheet.cell(row=i,column=1).value = str(resultSheet.cell(row=i,column=1).value) + " " + str(tempResult[k][0])
            continue
        elif(tempResult[k][1]=='MAG'):
            resultSheet.cell(row=i,column=1).value = str(resultSheet.cell(row=i,column=1).value) + " " + str(tempResult[k][0])
            continue
        elif(tempResult[k][1]=='XR'):
            resultSheet.cell(row=i,column=1).value = str(resultSheet.cell(row=i,column=1).value) + " " + str(tempResult[k][0])
            continue


for i in range(2,max_one_exception_row):
    if(exceptionSheet.cell(row=i,column=2).value==None):
        exceptionSheet.cell(row=i,column=2).value=""

for i in range(2,max_data_row):
    temp2 = resultSheet.cell(row=i,column=1).value.split(" ")
    for j in range(0,len(temp2)):
        for k in range(2,max_one_exception_row):
            if(temp2[j]==exceptionSheet.cell(row=k,column=1).value):
                temp2[j] = exceptionSheet.cell(row=k,column=2).value
    resultSheet.cell(row=i,column=1).value = ''
    for k in range(0, len(temp2)):
        if(temp2[k]!=''):
            resultSheet.cell(row=i,column=1).value = resultSheet.cell(row=i,column=1).value + temp2[k] +" "

result.save('result1.xlsx')