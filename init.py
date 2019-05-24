import openpyxl 
from konlpy.tag import Kkma
from konlpy.utils import pprint

kkma = Kkma()

result = openpyxl.Workbook()
resultSheet = result['Sheet']

test = openpyxl.load_workbook(filename='data.xlsx')
sheet =  test.worksheets[0]
exception = openpyxl.load_workbook(filename='exception.xlsx')
exceptionSheet = exception.worksheets[0]
max_one_exception_row = exceptionSheet.max_row + 1
max_two_exception_row = exceptionSheet.max_row + 1
max_data_row = sheet.max_row + 1

for i in range(2, max_data_row):
    if(sheet.cell(row = i, column = 1).value==None):
        max_data_row = i
        break
for i in range(2, max_one_exception_row):
    if(exceptionSheet.cell(row = i, column = 1).value==None):
        max_one_exception_row = i
        break
for i in range(2, max_two_exception_row):
    if(exceptionSheet.cell(row = i, column = 3).value==None):
        max_two_exception_row = i
        break

for i in range(2,max_data_row):
    for j in range(1,3):
        resultSheet.cell(row=i,column=j).value=""
        tempResult = kkma.pos(sheet.cell(row = i, column = j).value)
        index = 0
        for k in tempResult:
            if(tempResult[index][1]=='OL'):
                if(tempResult[index][0].lower()=='r'):
                    if(index+1<len(tempResult)):
                        if(tempResult[index+1][0]!='&'):
                            resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0].lower())
                    else:
                        resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0].lower())
                elif(tempResult[index][0].lower()!='d'):
                    resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0].lower())
            elif(tempResult[index][1]=='NNG'):
                resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0])
            elif(tempResult[index][1]=='NNP'):
                resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0])
            elif(tempResult[index][1]=='SW'):
                if(tempResult[index][0]=='++'):
                    resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + str(tempResult[index][0])
                elif(tempResult[index][0]=='#'):
                    resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + str(tempResult[index][0])
            elif(tempResult[index][1]=='NR'):
                if(index+1<len(tempResult)):
                    if(tempResult[index][0]=='3'):
                        if(tempResult[index+1][0].lower()=='d'):
                            resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0])+"d"
                    elif(tempResult[index][0]=='2'):
                        if(tempResult[index+1][0].lower()=='d'):
                            resultSheet.cell(row=i,column=j).value = str(resultSheet.cell(row=i,column=j).value) + " " + str(tempResult[index][0])+"d"
            index+=1

for i in range(2,max_data_row):
    temp = sheet.cell(row = i, column = 3).value.replace("\n#"," ")  
    temp = temp.replace('\xb7'," ")
    temp = temp.replace("("," ")
    temp = temp.replace(")"," ")
    temp = temp.replace("  "," ")
    temp = temp.replace("#","",1)
    resultSheet.cell(row=i,column=3).value= temp.lower()


for i in range(2,max_two_exception_row):
    for j in range(2,max_data_row):
        resultSheet.cell(row=j,column=1).value = resultSheet.cell(row=j,column=1).value.replace(exceptionSheet.cell(row=i,column=3).value,exceptionSheet.cell(row=i,column=4).value)
        resultSheet.cell(row=j,column=2).value = resultSheet.cell(row=j,column=2).value.replace(exceptionSheet.cell(row=i,column=3).value,exceptionSheet.cell(row=i,column=4).value)

for i in range(2,max_one_exception_row):
    if(exceptionSheet.cell(row=i,column=2).value==None):
        exceptionSheet.cell(row=i,column=2).value=""

for i in range(2,max_data_row):
    temp2 = resultSheet.cell(row=i,column=1).value.split(" ")
    for j in range(0,len(temp2)):
        for k in range(2,max_one_exception_row):
            if(temp2[j]==exceptionSheet.cell(row=k,column=1).value):
                temp2[j] = exceptionSheet.cell(row=k,column=2).value
    resultSheet.cell(row=i,column=1).value = ''
    for k in range(0, len(temp2)):
        if(temp2[k]!=''):
            resultSheet.cell(row=i,column=1).value = resultSheet.cell(row=i,column=1).value + temp2[k] +" "

    temp3 = resultSheet.cell(row=i,column=2).value.split(" ")
    for j in range(0,len(temp3)):
        for k in range(2,max_one_exception_row):
            if(temp3[j]==exceptionSheet.cell(row=k,column=1).value):
                temp3[j] = exceptionSheet.cell(row=k,column=2).value
    resultSheet.cell(row=i,column=2).value = ''
    for k in range(0, len(temp3)):
        if(temp3[k]!=''):
            resultSheet.cell(row=i,column=2).value = resultSheet.cell(row=i,column=2).value + temp3[k] +" "
    

result.save('result.xlsx')
