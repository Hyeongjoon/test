import openpyxl
from efficient_apriori import apriori
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('text', type=str,
                help="확인 하고싶은 단어")
args = parser.parse_args()

test = openpyxl.load_workbook(filename='words_result.xlsx')
resultSheet = test.worksheets[0]

max_data_row = resultSheet.max_row + 1

for i in range(2, max_data_row):
    if(resultSheet.cell(row = i, column = 1).value==None):
        max_data_row = i
        break

preData = []

for i in range(2,max_data_row):
    temp = resultSheet.cell(row=i,column=1).value.split(" ")
    tempData = []
    for k in range(0,len(temp)):
        if(temp[k]==''):
            continue
        else:
            tempData.append(temp[k])
    tempData = list(set(tempData))
    preData.append(tuple(tempData))

itemsets, rules = apriori(preData, min_support=0.10,  min_confidence=0.1)
dd = []
rules_rhs = filter(lambda rule: len(rule.lhs) == 1 and len(rule.rhs) == 1, rules)
for rule in sorted(rules_rhs, key=lambda rule: rule.confidence):
    if(rule.lhs[0]==args.text):
        for i in range(0, len(rule.rhs)):
            dd.append(rule.rhs[i])
dd.reverse()

print(dd)
