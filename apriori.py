import openpyxl
import argparse
from efficient_apriori import apriori

parser = argparse.ArgumentParser()
parser.add_argument('text', type=str,
                help="확인 하고싶은 단어")
args = parser.parse_args()

test = openpyxl.load_workbook(filename='result.xlsx')
resultSheet = test.worksheets[0]

exception = openpyxl.load_workbook(filename='exception.xlsx')
exceptionSheet = exception.worksheets[0]

max_data_row = resultSheet.max_row + 1
max_exceoption_row = exceptionSheet.max_row + 1

for i in range(2, max_exceoption_row):
    if(exceptionSheet.cell(row = i, column = 6).value==None):
        max_exceoption_row = i
        break


data = []
for i in range(2,max_data_row):
    data.append(str(resultSheet.cell(row=i,column=1).value) + str(resultSheet.cell(row=i,column=2).value) + str(resultSheet.cell(row=i,column=3).value))

preData = []
for j in range(0,len(data)):
    temp = data[j].replace("/", " ").split(" ")
    tempData = []
    for k in range(0,len(temp)):
        if(temp[k]==''):
            continue
        temp1 = temp[k].replace("*"," ")
        tempData.append(temp1)
    tempData = list(set(tempData))
    preData.append(tuple(tempData))


itemsets, rules = apriori(preData, min_support=0.15,  min_confidence=0.1)


dd = []
rules_rhs = filter(lambda rule: len(rule.lhs) == 1 and len(rule.rhs) == 1, rules)
for rule in sorted(rules_rhs, key=lambda rule: rule.confidence):
    if(rule.lhs[0]==args.text and rule.confidence!=1 and rule.lift>1 and rule.conviction < 1.5):
        for i in range(0, len(rule.rhs)):
            dd.append(rule.rhs[i])
dd.reverse()

for i in range(0,len(dd)):
    for j in range(2,max_exceoption_row):
        if(str(exceptionSheet.cell(row=j,column=6).value).find(dd[i])!=-1):
            tempStr = str(exceptionSheet.cell(row=j,column=6).value).split("/")
            dd[i] = tempStr[0]
            break


resultData = []
for i in range(0,len(dd)):
    check = False
    for j in range(0,len(resultData)):
        if(dd[i]==resultData[j]):
            check =True
    if(check==False):
        resultData.append(dd[i])


for i in range(0,len(resultData)):
    print(resultData[i])

